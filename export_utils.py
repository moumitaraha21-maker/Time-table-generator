from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from app import db, Timetable, TimetableEntry, Subject, Teacher, Classroom, Class

def export_to_pdf(timetable_id):
    """Export timetable to PDF"""
    timetable = Timetable.query.get(timetable_id)
    entries = TimetableEntry.query.filter_by(timetable_id=timetable_id).all()
    
    filename = f'timetable_{timetable_id}.pdf'
    filepath = os.path.join('exports', filename)
    os.makedirs('exports', exist_ok=True)
    
    doc = SimpleDocTemplate(filepath, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph(f"<b>{timetable.name}</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Get all classes
    classes = Class.query.all()
    
    for class_obj in classes:
        class_title = Paragraph(f"<b>Class {class_obj.grade} {class_obj.section}</b>", styles['Heading2'])
        elements.append(class_title)
        elements.append(Spacer(1, 10))
        
        # Create table data
        days = ['Period', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        data = [days]
        
        for period in range(8):
            row = [f'P{period+1}']
            for day in range(5):
                entry = TimetableEntry.query.filter_by(
                    timetable_id=timetable_id,
                    class_id=class_obj.id,
                    day_of_week=day,
                    period_number=period
                ).first()
                
                if entry:
                    subject = Subject.query.get(entry.subject_id)
                    teacher = Teacher.query.get(entry.teacher_id)
                    room = Classroom.query.get(entry.room_id)
                    cell_text = f"{subject.code}\n{room.name}"
                else:
                    cell_text = '-'
                
                row.append(cell_text)
            data.append(row)
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 30))
    
    doc.build(elements)
    return filepath

def export_to_excel(timetable_id):
    """Export timetable to Excel"""
    timetable = Timetable.query.get(timetable_id)
    entries = TimetableEntry.query.filter_by(timetable_id=timetable_id).all()
    
    filename = f'timetable_{timetable_id}.xlsx'
    filepath = os.path.join('exports', filename)
    os.makedirs('exports', exist_ok=True)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    classes = Class.query.all()
    days_name = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    
    for class_obj in classes:
        ws = wb.create_sheet(f"{class_obj.grade} {class_obj.section}")
        
        # Header
        ws['A1'] = 'Period'
        for col_idx, day in enumerate(days_name, start=2):
            ws.cell(row=1, column=col_idx).value = day
        
        # Style header
        for col in range(1, 7):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Fill periods
        for period in range(8):
            ws.cell(row=period+2, column=1).value = f'Period {period+1}'
            ws.cell(row=period+2, column=1).font = Font(bold=True)
            
            for day in range(5):
                entry = TimetableEntry.query.filter_by(
                    timetable_id=timetable_id,
                    class_id=class_obj.id,
                    day_of_week=day,
                    period_number=period
                ).first()
                
                if entry:
                    subject = Subject.query.get(entry.subject_id)
                    teacher = Teacher.query.get(entry.teacher_id)
                    room = Classroom.query.get(entry.room_id)
                    cell_value = f"{subject.name}\n{room.name}\n{entry.start_time}-{entry.end_time}"
                else:
                    cell_value = '-'
                
                cell = ws.cell(row=period+2, column=day+2)
                cell.value = cell_value
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Adjust column widths
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
    
    wb.save(filepath)
    return filepath
